#Definimos los módulos a importar
import pandas as pd
#Definimos las funciones a utilizar
from statistics import median,mean,mode,multimode
from math import isnan,floor
from itertools import filterfalse
import openpyxl
from tabulate import tabulate
#Definimos una serie simple numérica
serie=pd.Series([1,2,3])
print(serie)
print("La media es:"+str(median(serie)))

#Definimos una serie de tipo cadena o texto
s = pd.Series(['Matemáticas', 'Historia', 'Economía', 'Programación', 'Inglés'], dtype='string')
print(s)
serie.name='nombre'
#Ordenación de la serie a nivel texto o cadena
print(sorted(s))
#Definimos una serie con columnas
df=pd.DataFrame({'Columna 1':[1,2,3,4],'Columna 2':['a','b','c','d']})
#Esto es una prueba para actualizar
print(df)
data=[20.7,15,63.25,142.14]
data2=[10,52,10,63,87,63,10,63]
print(sorted(data))
print(sorted(data,key=None,reverse=True))
print(median(data)) #Para obtener la mediana
print(sum(data)) #Obtener la sumatoria
print(floor(mean(data))) #Para obtener la media con rendondeo
print(mode(data2))
print(multimode(data2))

#Obtenemos el tamaño de la tabla 
print(df.shape)
archivo='data/Resultados.xlsx'
excel_dataframe=openpyxl.load_workbook(archivo)
#Conocemos la hoja de calculo que se encuentra activa
print(excel_dataframe.active)
data_frame=excel_dataframe.active
data=[]
for row in range(1,data_frame.max_row):
    _row=[row]
    for col in data_frame.iter_cols(1,data_frame.max_column):
        #print(col[row].value)
        _row.append(col[row].value)
    data.append(_row)
headers=["Herramienta","Dominio del tema","Participación del grupo","Buena estructura","Claridad en contenido","Ejemplos","Tono de voz", "Tiempo de exposición","Imagenes"]
print(tabulate(data,headers=headers))